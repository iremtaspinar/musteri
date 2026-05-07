from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from .models import IncomingDocument, OutgoingDocument, Customer, Job
@login_required
def home(request): return render(request,'takip/home.html')
@login_required
def export_evrak_excel(request):
    wb=Workbook(); ws=wb.active; ws.title='Gelen Evrak'
    ws.append(['Tarih','Sayı','Geldiği Yer','Konu','Açıklama'])
    for r in IncomingDocument.objects.all().order_by('date','number'):
        ws.append([r.date,r.number,r.sender,r.subject,r.description])
    ws2=wb.create_sheet('Giden Evrak')
    ws2.append(['Tarih','Sayısı','Konu','Gönderildiği Yer','Açıklama'])
    for r in OutgoingDocument.objects.all().order_by('date','number'):
        ws2.append([r.date,r.number,r.subject,r.recipient,r.description])
    response=HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition']='attachment; filename="evrak_kayit.xlsx"'
    wb.save(response); return response
@login_required
def export_customer_payment_excel(request):
    wb=Workbook(); ws=wb.active; ws.title='Müşteri Ödeme Özeti'
    ws.append(['Müşteri','İş Sayısı','Toplam Ödenecek','Toplam Alınan','Toplam Kalan'])
    rows=[]
    for c in Customer.objects.all():
        jobs=Job.objects.filter(customer=c)
        if jobs.exists():
            total=sum([j.total for j in jobs]); paid=sum([j.paid for j in jobs]); remaining=sum([j.remaining for j in jobs])
            rows.append([c.name,jobs.count(),float(total),float(paid),float(remaining)])
    rows.sort(key=lambda x:x[4], reverse=True)
    for row in rows: ws.append(row)
    response=HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition']='attachment; filename="musteri_odeme_ozeti.xlsx"'
    wb.save(response); return response
